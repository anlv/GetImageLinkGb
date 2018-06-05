#!/usr/bin/python3

#!/usr/bin/python3
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from openpyxl.formula import Tokenizer


#read file excel using input
wb = load_workbook(sys.argv[1])
sheet = wb.worksheets[0]
print("Phai them cot ben phai")
xcolumn = int(input('Column: '))
xfrom = int(input ('from: '))
xto = int(input ('to: '))


for xrow in range(xfrom, xto):
    
    cell = sheet.cell(row = xrow, column= xcolumn)
    print(xrow)
    print(xcolumn)
    print(cell.value)
    #in 1 cell
    if 'HYPERLINK' in cell.value:
        tok = Tokenizer(cell.value)
        link = tok.items[1].value.replace('\"', '')
    else:
        link = cell.hyperlink.target


    print(link)
    
    result = requests.get(link)
    if result.status_code != 200:
        print('Error: ' + link )
        continue
    else:
        soup = BeautifulSoup(result.text, "html.parser")
        
        img = soup.find(id="product-image-large")
        if img is None:
            img = soup.find(id="product-image")

        img = img.find('img')
        srcf = img.get('src')
        srcb =  srcf.replace('front.png','back.png')
        sheet.cell(row = xrow, column= xcolumn +1).value = srcf + ', ' + srcb
    
# write link
wb.save(sys.argv[1])

#close file



