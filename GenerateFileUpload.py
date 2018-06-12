#!/usr/bin/python3
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
import requests
import re
from bs4 import BeautifulSoup
from openpyxl.formula import Tokenizer

#read file excel using input
wb = load_workbook(sys.argv[1])
sheetIndex = int(input("Nhap index cua sheet: "))
sheet = wb.worksheets[sheetIndex]

#xcolumn la cot chua link trang gear
rCodeC = 1
rLinkImageC = 4
rTitleC = 9

wCodeC = 1
wCodeTitleC = 2
wDescriptionC = 3
wTypeC = 5
wTagsC =  6
wSKUC = 14
wInventoryPolicyC  = 18
wFulfillmentC = 19
wPriceC = 20
wPriceCompareC = 21
wRequireShippingC = 22
wMainImageC = 25
wGiftCardC = 27 # value FALSE
wSubImageC = 43
wWeightUnitC = 44 #size (mug: oz)




xfrom = int(input ('from: '))
xto = int(input ('to: '))

pattern = r'([0-9]+)f(.)(([0-9]+)|(.))(.)'
## get group 1, 2, 3, 6

# read from row to row

#current write row
currentWRow = 1

for i in range(xfrom, xto):
    code = sheet.cell(row = i, column = rCodeC)
    print(code)
    if code.value is None:
        print("Code None")
        continue
    
    # extract code
    match = re.search(pattern, code.value)
    if match:
        idNumber = match.group(1)
        forObject = match.group(2)
        colorNumber = match.group(3)
        productType = match.group(6)
        print("\t" + idNumber + " " + forObject + " " + colorNumber + " " + productType)

        # check config file
        

    

        # get image link

        
    else:
        print("Code error: not match")
        continue


wb.save(sys.argv[1])
